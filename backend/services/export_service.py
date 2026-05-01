"""
Export service – generates Excel (.xlsx) files for invoices.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styling constants ────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_ACCENT_FILL = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
_ACCENT_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")

_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="0A2540")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="0A2540")
_NORMAL_FONT = Font(name="Calibri", size=11)
_MONEY_FORMAT = '#,##0.00'

_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _style_data_cell(cell, is_money=False):
    cell.font = _NORMAL_FONT
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(vertical="center")
    if is_money:
        cell.number_format = _MONEY_FORMAT


def _auto_width(ws, col_count, min_width=12):
    for col in range(1, col_count + 1):
        max_len = min_width
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(max_len, 40)


# ── Single invoice → Excel ───────────────────────────────────────────────────

def generate_single_invoice_excel(invoice_doc: dict) -> bytes:
    """Return .xlsx bytes for a single invoice."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    data = invoice_doc.get("data") or {}
    totals = data.get("totals") or {}
    bill_to = data.get("bill_to") or {}
    line_items = data.get("line_items") or []

    # ── Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "SmartInvoice AI"
    title_cell.font = Font(name="Calibri", bold=True, size=18, color="0A2540")

    ws.merge_cells("A2:F2")
    ws["A2"].value = "INVOICE"
    ws["A2"].font = Font(name="Calibri", bold=True, size=14, color="10B981")

    # ── Invoice details
    details = [
        ("Invoice #", data.get("invoice_no") or "N/A"),
        ("Date", data.get("date") or "N/A"),
        ("Due Date", data.get("due_date") or "N/A"),
        ("Vendor", data.get("vendor_name") or "N/A"),
        ("Currency", totals.get("currency") or "N/A"),
        ("Payment Method", data.get("payment_method") or "N/A"),
        ("Status", invoice_doc.get("status", "N/A").upper()),
    ]

    row = 4
    for label, value in details:
        ws.cell(row=row, column=1, value=label).font = _SUBTITLE_FONT
        ws.cell(row=row, column=2, value=value).font = _NORMAL_FONT
        row += 1

    # ── Bill To
    row += 1
    ws.cell(row=row, column=1, value="BILL TO").font = _SUBTITLE_FONT
    row += 1
    bill_details = [
        ("Name", bill_to.get("name") or "N/A"),
        ("Address", bill_to.get("address") or "N/A"),
        ("Email", bill_to.get("email") or "N/A"),
    ]
    for label, value in bill_details:
        ws.cell(row=row, column=1, value=label).font = _NORMAL_FONT
        ws.cell(row=row, column=2, value=value).font = _NORMAL_FONT
        row += 1

    # ── Line items table
    row += 1
    ws.cell(row=row, column=1, value="LINE ITEMS").font = _SUBTITLE_FONT
    row += 1

    headers = ["#", "Description", "Quantity", "Unit Price", "Total"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=header)
    _style_header_row(ws, row, len(headers))
    row += 1

    for i, item in enumerate(line_items, 1):
        ws.cell(row=row, column=1, value=i)
        _style_data_cell(ws.cell(row=row, column=1))

        ws.cell(row=row, column=2, value=item.get("description") or "")
        _style_data_cell(ws.cell(row=row, column=2))

        qty = item.get("quantity")
        ws.cell(row=row, column=3, value=qty if qty is not None else "")
        _style_data_cell(ws.cell(row=row, column=3))

        up = item.get("unit_price")
        ws.cell(row=row, column=4, value=up if up is not None else "")
        _style_data_cell(ws.cell(row=row, column=4), is_money=True)

        total = item.get("total")
        ws.cell(row=row, column=5, value=total if total is not None else "")
        _style_data_cell(ws.cell(row=row, column=5), is_money=True)

        row += 1

    # ── Totals
    row += 1
    totals_data = [
        ("Subtotal", totals.get("subtotal")),
        ("Tax", totals.get("tax")),
        ("Discount", totals.get("discount")),
    ]
    for label, value in totals_data:
        ws.cell(row=row, column=4, value=label).font = _NORMAL_FONT
        cell = ws.cell(row=row, column=5, value=value if value is not None else 0)
        _style_data_cell(cell, is_money=True)
        row += 1

    # Grand total row
    ws.cell(row=row, column=4, value="GRAND TOTAL").font = _ACCENT_FONT
    ws.cell(row=row, column=4).fill = _ACCENT_FILL
    gt = totals.get("grand_total")
    gt_cell = ws.cell(row=row, column=5, value=gt if gt is not None else 0)
    gt_cell.font = _ACCENT_FONT
    gt_cell.fill = _ACCENT_FILL
    gt_cell.number_format = _MONEY_FORMAT
    gt_cell.border = _THIN_BORDER

    # ── Notes
    if data.get("notes"):
        row += 2
        ws.cell(row=row, column=1, value="Notes:").font = _SUBTITLE_FONT
        row += 1
        ws.cell(row=row, column=1, value=data["notes"]).font = _NORMAL_FONT

    _auto_width(ws, 5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── All invoices → Excel ─────────────────────────────────────────────────────

def generate_all_invoices_excel(invoice_docs: list) -> bytes:
    """Return .xlsx bytes with a summary sheet + line-items sheet."""
    wb = Workbook()

    # ── Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Invoices Summary"

    summary_headers = [
        "Invoice #", "Vendor", "Date", "Due Date", "Bill To",
        "Subtotal", "Tax", "Discount", "Grand Total", "Currency",
        "Status", "Created At",
    ]

    for col_idx, header in enumerate(summary_headers, 1):
        ws_summary.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws_summary, 1, len(summary_headers))

    for row_idx, doc in enumerate(invoice_docs, 2):
        data = doc.get("data") or {}
        totals = data.get("totals") or {}
        bill_to = data.get("bill_to") or {}

        values = [
            data.get("invoice_no") or str(doc.get("_id", ""))[:8],
            data.get("vendor_name") or doc.get("original_filename") or "",
            data.get("date") or "",
            data.get("due_date") or "",
            bill_to.get("name") or "",
            totals.get("subtotal"),
            totals.get("tax"),
            totals.get("discount"),
            totals.get("grand_total"),
            totals.get("currency") or "",
            (doc.get("status") or "").upper(),
            doc.get("created_at") or "",
        ]

        money_cols = {6, 7, 8, 9}  # 1-indexed: subtotal, tax, discount, grand_total
        for col_idx, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            _style_data_cell(cell, is_money=(col_idx in money_cols))

    _auto_width(ws_summary, len(summary_headers))

    # ── Sheet 2: All Line Items
    ws_items = wb.create_sheet("Line Items")

    item_headers = [
        "Invoice #", "Vendor", "Item #", "Description",
        "Quantity", "Unit Price", "Total",
    ]
    for col_idx, header in enumerate(item_headers, 1):
        ws_items.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws_items, 1, len(item_headers))

    item_row = 2
    for doc in invoice_docs:
        data = doc.get("data") or {}
        inv_no = data.get("invoice_no") or str(doc.get("_id", ""))[:8]
        vendor = data.get("vendor_name") or ""

        for i, item in enumerate(data.get("line_items") or [], 1):
            values = [
                inv_no,
                vendor,
                i,
                item.get("description") or "",
                item.get("quantity"),
                item.get("unit_price"),
                item.get("total"),
            ]
            money_cols = {6, 7}
            for col_idx, value in enumerate(values, 1):
                cell = ws_items.cell(row=item_row, column=col_idx, value=value)
                _style_data_cell(cell, is_money=(col_idx in money_cols))
            item_row += 1

    _auto_width(ws_items, len(item_headers))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
